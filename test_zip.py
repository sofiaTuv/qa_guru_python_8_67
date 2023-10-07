import zipfile
from PyPDF2 import PdfReader
import xlrd
from openpyxl import load_workbook


def test_pdf():
    with zipfile.ZipFile('tmp/archive.zip', 'r') as myzip:
        with myzip.open('file_book.pdf') as pdf_t:
            pdf_t = PdfReader(pdf_t)
            page = pdf_t.pages[8]
            text = page.extract_text()
            assert 'Painting in Italy' in text

            number_of_page = len(pdf_t.pages)
            assert number_of_page == 256

            x = myzip.read('file_book.pdf')
            pdf_file_size = len(x)
            expected_file_size = 93212146
            assert pdf_file_size == expected_file_size


def test_txt():
    with zipfile.ZipFile('tmp/archive.zip', 'r') as myzip:
        with myzip.open('file_txt.txt') as txt_t:
            assert txt_t.read().decode('utf-8') == 'Test string 1 \r\nTest string 2 \r\nTest string 3'


def test_xls():
    with zipfile.ZipFile('tmp/archive.zip', 'r') as myzip:
        with myzip.open('file1.xls') as xls_t:
            xls_content = xls_t.read()
            xls_workbook = xlrd.open_workbook(file_contents=xls_content)
            sheet = xls_workbook.sheet_by_index(0)
            expected_value = '15/10/2017'
            actual_value = sheet.cell_value(1, 6)
            assert actual_value == expected_value


def test_xlsx():
    with zipfile.ZipFile('tmp/archive.zip', 'r') as myzip:
        with myzip.open('file2.xlsx') as xlsx_t:
            xlsx_file = load_workbook(xlsx_t)
            sheet = xlsx_file.active
            value = sheet.cell(row=3, column=3).value
            assert value == 'Hashimoto'
